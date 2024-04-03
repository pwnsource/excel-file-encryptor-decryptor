from cryptography.fernet import Fernet
from openpyxl import load_workbook

try:
    # Load the encryption key
    with open('encryption_key.txt', 'rb') as key_file:
        key = key_file.read()

    cipher_suite = Fernet(key)

    # Input filename
    encrypted_file_path = input("Enter the path of the encrypted Excel file: ")

    # Load the encrypted Excel file
    workbook = load_workbook(filename=encrypted_file_path)

    # Decrypt each cell in all sheets
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    cell.value = cipher_suite.decrypt(cell.value).decode()

    # Save the decrypted Excel file
    decrypted_file_path = input("Enter the name for the decrypted file (needs to end with .xlsx): ")
    workbook.save(decrypted_file_path)

except FileNotFoundError:
    print("File not found.")
