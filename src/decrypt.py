import os
import datetime
from cryptography.fernet import Fernet
from openpyxl import load_workbook

def log_error(error_message):
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_dir = os.path.join(os.path.dirname(__file__), '..', 'logs')
    log_file_path = os.path.join(log_dir, 'decrypter_log.txt')

    # Check if the logs directory exists, if not, create it
    if not os.path.exists(log_dir):
        os.makedirs(log_dir)

    with open(log_file_path, 'a') as log_file:
        log_file.write(f"[{timestamp}] {error_message}\n")

try:
    # Load the encryption key from the root folder
    key_path = os.path.join(os.path.dirname(__file__), '..', 'encryption_key.txt')
    with open(key_path, 'rb') as key_file:
        key = key_file.read()

    cipher_suite = Fernet(key)

    # Input filename
    encrypted_file_path = input("Enter the path of the encrypted Excel file: ")

    # Load the encrypted Excel file
    workbook = load_workbook(filename=encrypted_file_path)

    # Decrypt each cell in all sheets
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    cell.value = cipher_suite.decrypt(cell.value).decode()

    # Save the decrypted Excel file
    decrypted_file_path = input("Enter the name for the decrypted file (needs to end with .xlsx): ")
    workbook.save(decrypted_file_path)

except FileNotFoundError as e:
    error_message = f"File not found: {e}"
    print(error_message)
    log_error(error_message)

except Exception as e:
    error_message = f"An error occurred: {e}"
    print(error_message)
    log_error(error_message)
