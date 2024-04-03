import os
import datetime
from cryptography.fernet import Fernet
from openpyxl import load_workbook

def log_error(error_message):
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_dir = os.path.join(os.path.dirname(__file__), '..', 'logs')
    log_file_path = os.path.join(log_dir, 'encrypter_log.txt')

    # Check if the logs directory exists, if not, create it
    if not os.path.exists(log_dir):
        os.makedirs(log_dir)

    with open(log_file_path, 'a') as log_file:
        log_file.write(f"[{timestamp}] {error_message}\n")

try:
    # Load the encryption key
    with open(os.path.join(os.path.dirname(__file__), '..', 'encryption_key.txt'), 'rb') as key_file:
        key = key_file.read()

    cipher_suite = Fernet(key)

    # Input filename
    file_path = input("Enter the name of the original Excel file: ")

    # Load the Excel file
    workbook = load_workbook(filename=file_path)

    # Encrypt each cell in all sheets
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        for row in ws.iter_rows():
            for cell in row:
                cell.value = cipher_suite.encrypt(str(cell.value).encode())

    # Save the encrypted Excel file
    encrypted_file_path = input("Enter the name of the encrypted file (needs to end with .xlsx): ")
    workbook.save(encrypted_file_path)

    # Store the encryption key securely
    with open(os.path.join(os.path.dirname(__file__), '..', 'encryption_key.txt'), 'wb') as key_file:
        key_file.write(key)

except FileNotFoundError:
    error_message = "File not found."
    print(error_message)
    log_error(error_message)

except Exception as e:
    error_message = f"An error occurred: {e}"
    print(error_message)
    log_error(error_message)
